"""Importing the needed"""
import subprocess,os
import psutil
import tkinter as tk
import pyperclip
from selenium import webdriver
from selenium.common.exceptions import WebDriverException,NoSuchWindowException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import json
from openpyxl import Workbook
import random
from fake_useragent import UserAgent

"""Making a dynamic 'current working directory' variable to use"""
cwd = os.path.dirname(os.path.abspath(__file__))


"""GetDetailsFromWeb is the class that
is resposible for the data scrapping and execution"""
class GetDetailsFromWeb():
    """init module to initialize variables and call for the main method"""
    def __init__(self,needed_data_list):
        super().__init__()
        self.neended_data_list = needed_data_list
        self.xl_file_path = os.path.join(cwd,'home_details.xlsx')
        self.json_file_path = os.path.join(cwd,'home_details.json')
        print("needed data= ",self.neended_data_list)
        self.main()

    """get_chrome_path is able to fetch the chrome path input by the user"""
    def get_chrome_path(self):
        chrome_path=''
        if os.path.exists(os.path.join(cwd,"config.json")):
            with open (os.path.join(cwd,"config.json")) as f:
                data=json.load(f)
                chrome_path = data['chrome_path']
        
        return chrome_path

    """close_driver closes the driver and returns to the process"""
    def close_driver(self,driver):
        driver.close()
        return
    
    """quit_driver quits the driver and returns to the process"""
    def quit_driver(self,driver):
        driver.quit()
        return

    """is_chrome_window_closed is a special module that checks if the chrome browser is closed or not"""
    def is_chrome_window_closed(self):
        for window in psutil.process_iter(['pid', 'name']):
            try:
                if 'chrome' in window.info['name'].lower():
                    return False
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        return True

    """get_chrome_process is a module that fetches the chrome path from the user input,
    add some parameters to the script, and open the chrome window with a remote debugger attached to it"""
    def get_chrome_process(self):
        chrome_path = self.get_chrome_path()
            
        remote_debugging_port = 9222
        chrome_command = f'"{chrome_path}" --remote-debugging-port={remote_debugging_port}'

        if os.name == "posix":
            chrome_process = subprocess.Popen(chrome_command,shell=True)
        elif os.name == 'nt':
            chrome_process = subprocess.Popen(chrome_command)
            
        return chrome_process,remote_debugging_port

    """get_page_details is a module that does all the scrapping from the
    webpage according to the options selected by the user in the UI.py file"""
    def get_page_details(self,driver):
        time.sleep(1)
        features=[]
        data={}
        
        house_details=driver.find_elements(By.XPATH,"//div[@class='Inline__InlineContainer-sc-lf7x8d-0 gOaIBl View__InlineStyled-sc-4or7us-0 AYXFA property-info__property-attributes']")
        
        if 'house_name' in self.neended_data_list:
            house_name = driver.find_elements(By.XPATH, "//h1[@class='property-info-address']")
            if house_name:
                data["House Name"] = house_name[0].text
        
        if 'property_price' in self.neended_data_list:
            property_price=driver.find_elements(By.XPATH,"//span[@class='property-price property-info__price']")
            if property_price:
                data["Price"] = property_price[0].text
        
        if 'property_bonds' in self.neended_data_list:
            if 'rent' in driver.find_elements(By.XPATH,"(//a[@class='breadcrumb__link'])[1]")[0].get_attribute("href").lower():
                property_bonds=driver.find_elements(By.XPATH,"//div[@class='property-info__middle-content']//p[@class='Text__Typography-sc-vzn7fr-0 OdxXk']")
                if property_bonds:
                    data["Bonds"] = property_bonds[0].text
            else:
                data["Bonds"] = "Not applicable"
        
        if 'sold_date' in self.neended_data_list:
            if 'sold' in driver.find_elements(By.XPATH,"(//a[@class='breadcrumb__link'])[1]")[0].get_attribute("href").lower():
                property_sold=driver.find_elements(By.XPATH,"//div[@class='property-info__middle-content']//span")
                if property_sold:
                    for listings in property_sold:
                        if 'sold on' in listings.text.lower():
                            data["Sold On"] = listings.text
            else:
                data["Sold On"] = "Not applicable"
        
        if 'loan_repay_item' in self.neended_data_list:
            time.sleep(1)
            loan_repay_item=driver.find_elements(By.ID,"summary-repayments-container")
            if loan_repay_item:
                data["Loan Repay Amount"]=loan_repay_item[0].get_attribute("aria-label")
        
        if 'property_details' in self.neended_data_list:
            property_details=driver.find_elements(By.XPATH,"//span[@class='property-description__content']")
            if property_details:
                data["Description"]=str(property_details[0].text).replace("\\n", "\n")
        
        if 'property_features_item' in self.neended_data_list:
            property_features_item=driver.find_elements(By.XPATH,"//div[@class='property-features__feature']//p")
            if property_features_item:
                show_more_button=driver.find_elements(By.XPATH,"//button//p[contains(text(), 'Show more feature')]")
                if len(show_more_button) > 0:
                    show_more_button[0].click()
                for feature in property_features_item:
                    features.append(feature.text)
                features_value = ", ".join(features)
                if features_value:
                    data["Features"]=features_value
        
        if 'floorplan_area' in self.neended_data_list:
            floorplan_area=driver.find_elements(By.XPATH,"//span[@class='View__StyledIcon-sc-1am1guj-0 gPRTZR floorplans-tours__floorplan']")
            if floorplan_area:
                floorplan_area[0].click()
                time.sleep(1)
                floorplan_image=driver.find_elements(By.XPATH,"(//img[@class='pswp__img'])[1]")
                if floorplan_image:
                    data["Floor Plan"]=floorplan_image[0].get_attribute("src")
                    driver.find_element(By.XPATH,"//button[@title='Close (Esc)']").click()
        
        if 'agent_org' in self.neended_data_list:
            agent_org=driver.find_elements(By.XPATH,"(//a[@class='LinkBase-sc-12oy0hl-0 iskBYI sidebar-traffic-driver__name'])[last()]")
            if agent_org:
                data["Agent Organisation"]=agent_org[0].text
        
        if 'agent_org_address' in self.neended_data_list:
            agent_org_address=driver.find_elements(By.XPATH,"(//div[@class='sidebar-traffic-driver__detail-info'])[last()]")
            if agent_org_address:
                data["Agent Organisation Address"]=agent_org_address[0].text
        
        if 'property_size' in self.neended_data_list:
            property_size=house_details[0].find_elements(By.XPATH,"//div[@class='View__PropertySize-sc-1psmy31-0 cgTBlr property-size']")
            if property_size:
                data["Size"] = property_size[0].get_attribute("aria-label")
        
        if 'property_type' in self.neended_data_list:
            property_type=house_details[0].find_elements(By.XPATH,"//span[@class='property-info__property-type']")
            if property_type:
                data["Type"] = property_type[0].text
        
        if 'agent_phone' in self.neended_data_list:
            agent_phone=driver.find_elements(By.XPATH,"//ul[@class='agent-info agent-info--horizontal']/li//div[@class='phone']/a[1]")
            if agent_phone:
                if len(agent_phone) > 0:
                    names = []
                    for name_element in agent_phone:
                        names.append(str(name_element.get_attribute("href")).replace("tel:",''))
                    tel = "; ".join(set(names))
                data["Agent Phone"]=tel
        
        if 'agent_name' in self.neended_data_list:
            agent_name=driver.find_elements(By.XPATH,"//ul[@class='agent-info agent-info--horizontal']/li//a[contains(@class,'agent-info__name')]")
            if agent_name:
                if len(agent_name) > 1:
                    names = []
                    for name_element in agent_name:
                        names.append(name_element.text)
                    name = "; ".join(set(names))
                else:
                    name = agent_name[0].text
                data["Agent Name"] = name
        
        if 'house_properties_list' in self.neended_data_list:
            if house_details:
                house_properties_list=house_details[0].find_elements(By.XPATH,"//div[@class='View__PropertyDetail-sc-11ysrk6-0 haFtfe']")
                property_type=house_details[0].find_elements(By.XPATH,"//span[@class='property-info__property-type']")
                if house_properties_list:
                    for detail in house_properties_list:
                        try:
                            attribute = detail.get_attribute("aria-label")
                            if attribute:
                                if "bedroom" in attribute.lower():
                                    data["Bedrooms"] = attribute.title()
                                elif "bathroom" in attribute.lower():
                                    data["Bathroom"] = attribute.title()
                                elif "parking" in attribute.lower():
                                    data["Parking"] = attribute.title()
                                elif "study" in attribute.lower():
                                    data["Study"] = attribute.title()
                        except Exception as e:
                            print(f"Error: {e}")
        
        data["Property URL"]=driver.current_url

        return data
    
    """json_to_excel as the name says, it converts json file into the excel file with a special format"""
    def json_to_excel(self, json_file_path, excel_file_path):
        try:
            # Read the JSON data from the file
            with open(json_file_path, 'r', encoding='utf-8') as jsonfile:
                data = json.load(jsonfile)

            # Check if the data is a dictionary
            if not isinstance(data, dict):
                raise ValueError("JSON data must be a dictionary with categories as keys")

            # Create a new Excel workbook
            workbook = Workbook()
            # Remove the default sheet created with the workbook
            workbook.remove(workbook.active)

            # Iterate over each category in the JSON data
            for category, home_details in data.items():
                # Create a new sheet for the category
                sheet = workbook.create_sheet(title=category)

                # Filter out empty dictionaries or dictionaries with all empty values
                filtered_home_details = [home for home in home_details if home and any(home.values())]

                if not filtered_home_details:
                    continue  # Skip empty categories

                # Collect all unique fieldnames from the dictionaries
                fieldnames = set()
                for home in filtered_home_details:
                    fieldnames.update(home.keys())

                fieldnames = sorted(fieldnames)  # Sort fieldnames for consistent ordering

                # Write the header row
                for col_num, fieldname in enumerate(fieldnames, 1):
                    sheet.cell(row=1, column=col_num, value=fieldname)

                # Write the data rows
                for row_num, home in enumerate(filtered_home_details, start=2):
                    for col_num, fieldname in enumerate(fieldnames, 1):
                        sheet.cell(row=row_num, column=col_num, value=home.get(fieldname, "Not Available"))

            # Save the workbook to the specified file path
            workbook.save(excel_file_path)
        except FileNotFoundError:
            print("Json File does not exists")
            return
        except PermissionError:
            print("Permision error, retry after closing the files")
            return
        except Exception as e:
            print("Other Error: ",e)

    
    """write_home_details is the module that takes the scrapped data, and write it to the json file
    that can be used to create a excel file afterwards"""
    def write_home_details(self, category, home_details, json_file_path):
        category=category.title()
        try:
            # Initialize an empty dictionary for existing data
            existing_data = {}

            # Check if the JSON file exists and read the existing data if it does
            if os.path.isfile(json_file_path):
                with open(json_file_path, 'r', encoding='utf-8') as jsonfile:
                    try:
                        existing_data = json.load(jsonfile)
                    except json.JSONDecodeError:
                        existing_data = {}

            # If the category is not already in existing_data, initialize it as an empty list
            if category not in existing_data:
                existing_data[category] = []

            # Append new home details to the corresponding category in the existing data
            existing_data[category].extend(home_details)

            # Write the updated data back to the JSON file
            with open(json_file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(existing_data, jsonfile, ensure_ascii=False, indent=4)

        except PermissionError:
            # Handle permission error, usually occurs if the file is open elsewhere
            new_json_file_path = os.path.join(os.path.dirname(json_file_path), 'home_details_1.json')
            with open(new_json_file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(existing_data, jsonfile, ensure_ascii=False, indent=4)

    
    """get_random_useragent uses a module from fake_agent,
    that provides a random user agent to the process created.
    This module is used to try to bypass the HTTP error 429 ( Too Many Requests )"""
    def get_random_useragent(self):
        ua = UserAgent()
        chrome_user_agent = ua.chrome
        return chrome_user_agent

    
    """get_urls is the module that scrapps first 25 URLs from the page and passes it to the scrapper"""
    def get_urls(self,driver):
        data_dict={}
        house_names = driver.find_elements(By.XPATH, "//a[@class='details-link residential-card__details-link']//span")
        house_url = driver.find_elements(By.XPATH, "//a[@class='details-link residential-card__details-link']")
        for names,urls in zip(house_names,house_url):
            data_dict[urls.get_attribute("href")]=names.get_attribute('innerHTML')
        return(data_dict)

    """show_info_with_copy_button is a minute module for the Tkinter"""
    def show_info_with_copy_button(self,master_url):
        def on_ok():
            root.destroy()

        def on_copy_url():
            pyperclip.copy(master_url)
            root.destroy()

        root = tk.Tk()
        root.title("Info")
        root.attributes("-topmost", True)

        label = tk.Label(root, text=f"Please visit the URL: {master_url} and search for the location", wraplength=400)
        label.pack(pady=10)

        button_frame = tk.Frame(root)
        button_frame.pack(pady=10)

        ok_button = tk.Button(button_frame, text="OK", command=on_ok)
        ok_button.grid(row=0, column=0, padx=5)

        copy_button = tk.Button(button_frame, text="Copy URL", command=on_copy_url)
        copy_button.grid(row=0, column=1, padx=5)

        root.wait_window(root)
    
    """main is the main module that consists of the scrapped data.
    It manages the work flow. It calls the existing modules and magages the file and the data"""
    def main(self):
        master_url="https://www.realestate.com.au/"
        chrome_process,port=self.get_chrome_process()

        # Create Chrome options
        chrome_user_agent = self.get_random_useragent()
        chrome_options = Options()
        chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
        chrome_options.add_argument(f"user-agent={chrome_user_agent}")

        # Connect to the existing Chrome instance
        driver = webdriver.Chrome(options=chrome_options)

        try:
            time.sleep(1)
            driver.maximize_window()
        except WebDriverException as e:
            print(f"Error maximizing window: {e}")

        # Now you can interact with the existing Chrome session
        self.show_info_with_copy_button(master_url)
        original_url=''
        try:                                                                                       
            while not self.is_chrome_window_closed():
                while driver is not None and not any(substring in driver.current_url.lower() for substring in ["realestate.com.au/buy/", "realestate.com.au/rent/", "realestate.com.au/sold/"]):
                    if self.is_chrome_window_closed():
                        raise NoSuchWindowException
                    time.sleep(1)
                if driver.current_url != original_url:
                    original_url=driver.current_url
                    original_window = driver.current_window_handle
                    # If the URL contains '/buy/', fetch the data
                    data_dict = self.get_urls(driver=driver)
                    time.sleep(1)
                    try:
                        driver.switch_to.new_window('tab')
                        time.sleep(1)
                        home_details = []
                        counter=0
                        for url,data_list in data_dict.items():
                            counter+=1
                            driver.get(url)
                            time.sleep(1)
                            category=driver.find_elements(By.XPATH,"(//a[@class='breadcrumb__link'])[1]")[0].get_attribute("title")
                            names = self.get_page_details(driver)
                            time.sleep(random.uniform(9,15))
                            home_details.append(names)
                        self.write_home_details(category,home_details, self.json_file_path)
                        driver.close()
                        time.sleep(1)
                        driver.switch_to.window(original_window)
                    except Exception as e:
                        print(f"Error initializing headless WebDriver: {e}")


        except NoSuchWindowException:
            # Close the browser
            if driver:
                self.quit_driver(driver)
            
            # Terminate the Chrome process opened by the script
            if not self.is_chrome_window_closed():
                chrome_process.kill()
                chrome_process.wait()
            print("Selenium window closed")
        
        except AttributeError:
            # Close the browser
            if driver:
                self.quit_driver(driver)
            
            # Terminate the Chrome process opened by the script
            if not self.is_chrome_window_closed():
                chrome_process.kill()
                chrome_process.wait()
            print("Selenium window closed")
            
        
        finally:
            # Close the browser
            if driver:
                self.quit_driver(driver)
            
            # Terminate the Chrome process opened by the script
            if not self.is_chrome_window_closed():
                chrome_process.kill()
                chrome_process.wait()
            self.json_to_excel(self.json_file_path,self.xl_file_path)